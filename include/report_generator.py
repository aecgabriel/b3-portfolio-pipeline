"""
B3 Portfolio Report Generator
Generates a comprehensive Excel report from DuckDB mart tables.
Includes performance benchmarks (CDI, IBOV, DIVO11) and charts.
"""

import os
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, PieChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList

log = logging.getLogger(__name__)

# ── Style constants ──
DARK_BLUE = '1F4E79'
MED_BLUE = '2E75B6'
LIGHT_BLUE = 'D6E4F0'
VERY_LIGHT_BLUE = 'F2F7FB'
GREEN = '00B050'
RED = 'FF0000'
DARK_GREEN = '006100'
LIGHT_GREEN = 'C6EFCE'
DARK_RED = '9C0006'
LIGHT_RED = 'FFC7CE'
GRAY = 'F2F2F2'
WHITE = 'FFFFFF'

HEADER_FILL = PatternFill('solid', fgColor=DARK_BLUE)
HEADER_FONT = Font(name='Arial', bold=True, color=WHITE, size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor=MED_BLUE)
SUBHEADER_FONT = Font(name='Arial', bold=True, color=WHITE, size=10)
DATA_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True, color=DARK_BLUE)
SUBTITLE_FONT = Font(name='Arial', size=11, bold=True, color=MED_BLUE)
TOTAL_FILL = PatternFill('solid', fgColor=LIGHT_BLUE)
TOTAL_FONT = Font(name='Arial', size=10, bold=True)
ALT_FILL = PatternFill('solid', fgColor=VERY_LIGHT_BLUE)
GAIN_FONT = Font(name='Arial', size=10, color=DARK_GREEN)
LOSS_FONT = Font(name='Arial', size=10, color=DARK_RED)
GAIN_FILL = PatternFill('solid', fgColor=LIGHT_GREEN)
LOSS_FILL = PatternFill('solid', fgColor=LIGHT_RED)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
MONEY_FMT = '#,##0.00'
QTY_FMT = '#,##0'
PCT_FMT = '0.00%'
PCT2_FMT = '0.00"%"'

MONTH_LABELS = {
    '01': 'Jan', '02': 'Fev', '03': 'Mar', '04': 'Abr',
    '05': 'Mai', '06': 'Jun', '07': 'Jul', '08': 'Ago',
    '09': 'Set', '10': 'Out', '11': 'Nov', '12': 'Dez',
}


def _period_to_label(period: str) -> str:
    parts = period.split('-')
    if len(parts) == 2:
        return f"{MONTH_LABELS.get(parts[1], parts[1])}/{parts[0][2:]}"
    return period


def _period_to_date(period: str) -> datetime:
    parts = period.split('-')
    return datetime(int(parts[0]), int(parts[1]), 28)


def _apply_header(ws, row, headers, col_widths, start_col=1):
    for c, (h, w) in enumerate(zip(headers, col_widths), start_col):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 28


def _apply_data_cell(ws, row, col, value, fmt=None, align='right'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal=align)
    if row % 2 == 0:
        cell.fill = ALT_FILL
    if fmt:
        cell.number_format = fmt
    return cell


def _apply_gain_loss_style(cell, value):
    if value is None:
        return
    if value > 0:
        cell.font = GAIN_FONT
    elif value < 0:
        cell.font = LOSS_FONT


def _add_title(ws, row, title, subtitle=None):
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.row_dimensions[row].height = 24
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle).font = SUBTITLE_FONT
        return row + 3
    return row + 2


# ════════════════════════════════════════════════════════════════
# Data fetching
# ════════════════════════════════════════════════════════════════

def fetch_all_data(duckdb_path: str) -> Dict[str, pd.DataFrame]:
    con = duckdb.connect(duckdb_path, read_only=True)
    data = {}

    data['monthly_wallet'] = con.execute(
        "SELECT * FROM fct_monthly_wallet ORDER BY _reference_period"
    ).fetchdf()

    data['stock_detail'] = con.execute(
        "SELECT * FROM fct_stock_detail ORDER BY _reference_period, ticker"
    ).fetchdf()

    data['dividends'] = con.execute(
        "SELECT * FROM fct_dividend_summary ORDER BY _reference_period, ticker"
    ).fetchdf()

    data['treasury'] = con.execute(
        "SELECT * FROM fct_treasury_summary ORDER BY _reference_period, product_name"
    ).fetchdf()

    data['fixed_income'] = con.execute(
        "SELECT * FROM fct_fixed_income_summary ORDER BY _reference_period, product_name"
    ).fetchdf()

    data['trades'] = con.execute("""
        SELECT ticker, _reference_period, institution,
               buy_quantity, avg_buy_price, buy_quantity * avg_buy_price as total_buy_cost,
               sell_quantity, avg_sell_price, sell_quantity * avg_sell_price as total_sell_value
        FROM stg_trades
        ORDER BY _reference_period, ticker
    """).fetchdf()

    data['weighted_avg'] = con.execute(
        "SELECT * FROM int_weighted_avg_price ORDER BY ticker"
    ).fetchdf()

    con.close()
    return data


def fetch_benchmarks(periods: List[str]) -> pd.DataFrame:
    """Fetch CDI, IBOV, and DIVO11 benchmark data for comparison."""
    import requests

    dates = sorted([_period_to_date(p) for p in periods])
    start_date = dates[0].replace(day=1)
    end_date = dates[-1]

    benchmarks = pd.DataFrame({'_reference_period': periods})

    # ── CDI from BCB API (series 4391 = daily CDI rate) ──
    try:
        bcb_start = start_date.strftime('%d/%m/%Y')
        bcb_end = end_date.strftime('%d/%m/%Y')
        url = (
            f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.4391/dados"
            f"?formato=json&dataInicial={bcb_start}&dataFinal={bcb_end}"
        )
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200:
            cdi_data = resp.json()
            cdi_df = pd.DataFrame(cdi_data)
            cdi_df['data'] = pd.to_datetime(cdi_df['data'], format='%d/%m/%Y')
            cdi_df['valor'] = cdi_df['valor'].astype(float)
            cdi_df['year_month'] = cdi_df['data'].dt.strftime('%Y-%m')

            monthly_cdi = cdi_df.groupby('year_month').apply(
                lambda x: ((1 + x['valor'] / 100).prod() - 1) * 100,
                include_groups=False
            ).reset_index()
            monthly_cdi.columns = ['_reference_period', 'cdi_monthly_return']

            benchmarks = benchmarks.merge(monthly_cdi, on='_reference_period', how='left')
            benchmarks['cdi_monthly_return'] = benchmarks['cdi_monthly_return'].fillna(0)
            benchmarks['cdi_cumulative'] = (1 + benchmarks['cdi_monthly_return'] / 100).cumprod()
            benchmarks['cdi_cumulative'] = (benchmarks['cdi_cumulative'] - 1) * 100
            log.info("CDI data fetched successfully")
        else:
            log.warning(f"BCB API returned status {resp.status_code}")
            benchmarks['cdi_monthly_return'] = 0
            benchmarks['cdi_cumulative'] = 0
    except Exception as e:
        log.warning(f"Failed to fetch CDI data: {e}")
        benchmarks['cdi_monthly_return'] = 0
        benchmarks['cdi_cumulative'] = 0

    # ── IBOV and DIVO11 from yfinance ──
    try:
        import yfinance as yf

        yf_start = start_date.strftime('%Y-%m-%d')
        yf_end = (end_date + timedelta(days=5)).strftime('%Y-%m-%d')

        for ticker, col_name in [("^BVSP", "ibov"), ("DIVO11.SA", "divo11")]:
            try:
                hist = yf.download(ticker, start=yf_start, end=yf_end,
                                   interval="1mo", progress=False, auto_adjust=True)
                if hist.empty:
                    hist = yf.download(ticker, start=yf_start, end=yf_end,
                                       interval="1d", progress=False, auto_adjust=True)

                if not hist.empty:
                    if hasattr(hist.columns, 'levels'):
                        hist.columns = hist.columns.get_level_values(0)

                    hist = hist.reset_index()
                    date_col = 'Date' if 'Date' in hist.columns else hist.columns[0]
                    hist['year_month'] = pd.to_datetime(hist[date_col]).dt.strftime('%Y-%m')

                    monthly = hist.groupby('year_month').last().reset_index()
                    close_col = 'Close' if 'Close' in monthly.columns else 'Adj Close'

                    monthly[f'{col_name}_close'] = monthly[close_col].astype(float)
                    monthly = monthly[['year_month', f'{col_name}_close']].copy()
                    monthly.columns = ['_reference_period', f'{col_name}_close']

                    first_val = monthly[f'{col_name}_close'].iloc[0]
                    monthly[f'{col_name}_return'] = (
                        (monthly[f'{col_name}_close'] / first_val) - 1
                    ) * 100

                    benchmarks = benchmarks.merge(
                        monthly[['_reference_period', f'{col_name}_return']],
                        on='_reference_period', how='left'
                    )
                    benchmarks[f'{col_name}_return'] = benchmarks[f'{col_name}_return'].fillna(method='ffill').fillna(0)
                    log.info(f"{ticker} data fetched successfully")
                else:
                    benchmarks[f'{col_name}_return'] = 0
                    log.warning(f"No data returned for {ticker}")
            except Exception as e:
                log.warning(f"Failed to fetch {ticker}: {e}")
                benchmarks[f'{col_name}_return'] = 0

    except ImportError:
        log.warning("yfinance not installed, skipping IBOV/DIVO11 benchmarks")
        benchmarks['ibov_return'] = 0
        benchmarks['divo11_return'] = 0

    return benchmarks


# ════════════════════════════════════════════════════════════════
# Sheet builders
# ════════════════════════════════════════════════════════════════

def build_resumo(wb: Workbook, data: Dict, benchmarks: pd.DataFrame):
    """Sheet 1: Dashboard / Resumo Geral"""
    ws = wb.active
    ws.title = 'Resumo'
    ws.sheet_properties.tabColor = DARK_BLUE

    wallet = data['monthly_wallet']
    if wallet.empty:
        ws.cell(row=1, column=1, value="Sem dados disponíveis")
        return

    latest = wallet.iloc[-1]
    latest_period = latest['_reference_period']

    row = _add_title(ws, 1, 'Relatório de Carteira B3', f'Posição em {_period_to_label(latest_period)}')

    # ── Key metrics ──
    metrics = [
        ('Valor Total da Carteira', latest['total_portfolio_value'], MONEY_FMT),
        ('Total Investido', latest['total_invested'], MONEY_FMT),
        ('Ganho/Perda', latest['absolute_gain_loss'], MONEY_FMT),
        ('Retorno %', latest['pct_return'] / 100, PCT_FMT),
    ]

    ws.cell(row=row, column=1, value='Indicadores Principais').font = SUBTITLE_FONT
    row += 1
    for label, value, fmt in metrics:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.column_dimensions['A'].width = 28
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        cell.font = BOLD_FONT
        ws.column_dimensions['B'].width = 20
        if label == 'Ganho/Perda':
            _apply_gain_loss_style(cell, value)
        if label == 'Retorno %':
            _apply_gain_loss_style(cell, value)
        cell.border = THIN_BORDER
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1

    row += 1

    # ── Allocation by asset class ──
    ws.cell(row=row, column=1, value='Alocação por Classe de Ativo').font = SUBTITLE_FONT
    row += 1
    alloc_start = row
    classes = [
        ('Ações & ETFs', latest.get('equity_value', 0)),
        ('Tesouro Direto', latest.get('treasury_value', 0)),
        ('Renda Fixa', latest.get('fixed_income_value', 0)),
    ]
    total_val = sum(v for _, v in classes)

    for label, value in classes:
        ws.cell(row=row, column=1, value=label).font = DATA_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = MONEY_FMT
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        pct_cell = ws.cell(row=row, column=3, value=value / total_val if total_val > 0 else 0)
        pct_cell.number_format = PCT_FMT
        pct_cell.font = DATA_FONT
        pct_cell.border = THIN_BORDER
        ws.column_dimensions['C'].width = 12
        row += 1

    # Pie chart
    pie = PieChart()
    pie.title = "Alocação por Classe"
    pie.style = 10
    pie.width = 14
    pie.height = 10
    labels = Reference(ws, min_col=1, min_row=alloc_start, max_row=alloc_start + len(classes) - 1)
    values = Reference(ws, min_col=2, min_row=alloc_start, max_row=alloc_start + len(classes) - 1)
    pie.add_data(values, titles_from_data=False)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    ws.add_chart(pie, f"E{alloc_start - 1}")

    row += 2

    # ── Dividend summary ──
    divs = data['dividends']
    if not divs.empty:
        ws.cell(row=row, column=1, value='Proventos Recebidos (Total)').font = SUBTITLE_FONT
        row += 1
        div_by_type = divs.groupby('event_type')['total_income'].sum()
        for evt, val in div_by_type.items():
            ws.cell(row=row, column=1, value=str(evt)).font = DATA_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            cell = ws.cell(row=row, column=2, value=val)
            cell.number_format = MONEY_FMT
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            row += 1
        ws.cell(row=row, column=1, value='Total Proventos').font = TOTAL_FONT
        ws.cell(row=row, column=1).fill = TOTAL_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        total_cell = ws.cell(row=row, column=2, value=divs['total_income'].sum())
        total_cell.number_format = MONEY_FMT
        total_cell.font = TOTAL_FONT
        total_cell.fill = TOTAL_FILL
        total_cell.border = THIN_BORDER


def build_evolucao(wb: Workbook, data: Dict, benchmarks: pd.DataFrame):
    """Sheet 2: Evolução Patrimonial + Performance vs Benchmarks"""
    ws = wb.create_sheet('Evolução & Performance')
    ws.sheet_properties.tabColor = MED_BLUE

    wallet = data['monthly_wallet']
    if wallet.empty:
        return

    row = _add_title(ws, 1, 'Evolução Patrimonial Mensal')

    # ── Monthly data table ──
    headers = ['Período', 'Valor Carteira', 'Ações & ETFs', 'Tesouro', 'Renda Fixa',
               'Investido', 'Ganho/Perda', 'Retorno %']
    widths = [12, 18, 16, 16, 16, 18, 16, 12]
    _apply_header(ws, row, headers, widths)
    row += 1
    data_start = row

    for _, r in wallet.iterrows():
        _apply_data_cell(ws, row, 1, _period_to_label(r['_reference_period']), align='center')
        _apply_data_cell(ws, row, 2, r['total_portfolio_value'], MONEY_FMT)
        _apply_data_cell(ws, row, 3, r['equity_value'], MONEY_FMT)
        _apply_data_cell(ws, row, 4, r['treasury_value'], MONEY_FMT)
        _apply_data_cell(ws, row, 5, r['fixed_income_value'], MONEY_FMT)
        _apply_data_cell(ws, row, 6, r['total_invested'], MONEY_FMT)
        gl_cell = _apply_data_cell(ws, row, 7, r['absolute_gain_loss'], MONEY_FMT)
        _apply_gain_loss_style(gl_cell, r['absolute_gain_loss'])
        pct_cell = _apply_data_cell(ws, row, 8, r['pct_return'] / 100, PCT_FMT)
        _apply_gain_loss_style(pct_cell, r['pct_return'])
        row += 1

    data_end = row - 1

    # ── Portfolio value chart ──
    chart = LineChart()
    chart.title = "Evolução do Patrimônio"
    chart.style = 10
    chart.width = 28
    chart.height = 14
    chart.y_axis.title = "Valor (R$)"
    chart.x_axis.title = "Período"

    cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)

    for col_idx, name, color in [
        (2, "Valor Total", "1F4E79"),
        (3, "Ações & ETFs", "2E75B6"),
        (4, "Tesouro", "00B050"),
        (5, "Renda Fixa", "FFC000"),
    ]:
        vals = Reference(ws, min_col=col_idx, min_row=data_start - 1, max_row=data_end)
        chart.add_data(vals, titles_from_data=True)
        chart.series[-1].graphicalProperties.line.width = 25000
        chart.series[-1].graphicalProperties.line.solidFill = color

    chart.set_categories(cats)
    ws.add_chart(chart, f"A{row + 1}")
    row += 18

    # ── Performance vs Benchmarks ──
    row += 1
    row = _add_title(ws, row, 'Performance vs Benchmarks')

    bench_headers = ['Período', 'Carteira %', 'CDI %', 'IBOV %', 'DIVO11 %']
    bench_widths = [12, 14, 14, 14, 14]
    _apply_header(ws, row, bench_headers, bench_widths)
    row += 1
    bench_start = row

    # Calculate portfolio cumulative return
    first_invested = wallet.iloc[0]['total_invested']
    if first_invested <= 0:
        first_invested = wallet.iloc[0]['total_portfolio_value']

    for i, (_, w) in enumerate(wallet.iterrows()):
        period = w['_reference_period']
        _apply_data_cell(ws, row, 1, _period_to_label(period), align='center')

        portfolio_cum = w['pct_return']
        _apply_data_cell(ws, row, 2, portfolio_cum / 100, PCT_FMT)

        bench_row = benchmarks[benchmarks['_reference_period'] == period]
        if not bench_row.empty:
            br = bench_row.iloc[0]
            _apply_data_cell(ws, row, 3, br.get('cdi_cumulative', 0) / 100, PCT_FMT)
            _apply_data_cell(ws, row, 4, br.get('ibov_return', 0) / 100, PCT_FMT)
            _apply_data_cell(ws, row, 5, br.get('divo11_return', 0) / 100, PCT_FMT)
        else:
            for c in range(3, 6):
                _apply_data_cell(ws, row, c, 0, PCT_FMT)
        row += 1

    bench_end = row - 1

    # Benchmark chart
    bench_chart = LineChart()
    bench_chart.title = "Retorno Acumulado: Carteira vs CDI vs IBOV vs DIVO11"
    bench_chart.style = 10
    bench_chart.width = 28
    bench_chart.height = 14
    bench_chart.y_axis.title = "Retorno Acumulado (%)"
    bench_chart.y_axis.numFmt = '0.00%'

    cats = Reference(ws, min_col=1, min_row=bench_start, max_row=bench_end)

    colors = ["1F4E79", "FF6600", "00B050", "7030A0"]
    for col_idx in range(2, 6):
        vals = Reference(ws, min_col=col_idx, min_row=bench_start - 1, max_row=bench_end)
        bench_chart.add_data(vals, titles_from_data=True)
        bench_chart.series[-1].graphicalProperties.line.width = 25000
        bench_chart.series[-1].graphicalProperties.line.solidFill = colors[col_idx - 2]

    bench_chart.set_categories(cats)
    ws.add_chart(bench_chart, f"A{row + 1}")


def build_detalhamento_acoes(wb: Workbook, data: Dict):
    """Sheet 3: Detalhamento Ações & ETFs (última posição + histórico)"""
    ws = wb.create_sheet('Ações & ETFs')
    ws.sheet_properties.tabColor = '2E75B6'

    detail = data['stock_detail']
    if detail.empty:
        return

    latest_period = detail['_reference_period'].max()
    latest = detail[detail['_reference_period'] == latest_period].copy()

    row = _add_title(ws, 1, f'Detalhamento de Ações & ETFs',
                     f'Posição em {_period_to_label(latest_period)}')

    headers = ['Ticker', 'Tipo', 'Quantidade', 'Preço Médio', 'Preço Tela',
               'Valor Investido', 'Valor Tela', 'Lucro/Prejuízo', 'Variação %']
    widths = [12, 10, 12, 16, 16, 18, 18, 18, 14]
    _apply_header(ws, row, headers, widths)
    row += 1
    data_start = row

    for _, s in latest.iterrows():
        _apply_data_cell(ws, row, 1, s['ticker'], align='left')
        _apply_data_cell(ws, row, 2, 'ETF' if s['asset_type'] == 'etf' else 'Ação', align='center')
        _apply_data_cell(ws, row, 3, s['total_quantity'], QTY_FMT)
        _apply_data_cell(ws, row, 4, s['mean_purchase_price'], MONEY_FMT)
        _apply_data_cell(ws, row, 5, s['current_closing_price'], MONEY_FMT)
        _apply_data_cell(ws, row, 6, s['total_invested'], MONEY_FMT)
        _apply_data_cell(ws, row, 7, s['total_current_value'], MONEY_FMT)

        gl_cell = _apply_data_cell(ws, row, 8, s['unrealized_gain_loss'], MONEY_FMT)
        _apply_gain_loss_style(gl_cell, s['unrealized_gain_loss'])

        pct_val = s['pct_price_change'] / 100 if s['pct_price_change'] else 0
        pct_cell = _apply_data_cell(ws, row, 9, pct_val, PCT_FMT)
        _apply_gain_loss_style(pct_cell, pct_val)
        row += 1

    # Total row
    for c in range(1, 10):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.font = TOTAL_FONT
    ws.cell(row=row, column=1, value='TOTAL')
    for c, letter in [(3, 'C'), (6, 'F'), (7, 'G'), (8, 'H')]:
        ws.cell(row=row, column=c, value=f'=SUM({letter}{data_start}:{letter}{row - 1})').number_format = MONEY_FMT if c > 3 else QTY_FMT

    ws.auto_filter.ref = f'A{data_start - 1}:I{row - 1}'
    ws.freeze_panes = f'A{data_start}'

    row += 2

    # ── Bar chart: Lucro/Prejuízo por ticker ──
    bar = BarChart()
    bar.type = "col"
    bar.title = "Lucro/Prejuízo por Ativo"
    bar.style = 10
    bar.width = 24
    bar.height = 12
    bar.y_axis.title = "R$"

    cats = Reference(ws, min_col=1, min_row=data_start, max_row=row - 3)
    vals = Reference(ws, min_col=8, min_row=data_start - 1, max_row=row - 3)
    bar.add_data(vals, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = "2E75B6"
    ws.add_chart(bar, f"A{row + 1}")


def build_proventos(wb: Workbook, data: Dict):
    """Sheet 4: Proventos Recebidos (IR)"""
    ws = wb.create_sheet('Proventos')
    ws.sheet_properties.tabColor = '00B050'

    divs = data['dividends']
    if divs.empty:
        ws.cell(row=1, column=1, value="Sem proventos registrados").font = DATA_FONT
        return

    row = _add_title(ws, 1, 'Proventos Recebidos', 'Dividendos, JCP e Rendimentos')

    # ── Summary by type ──
    ws.cell(row=row, column=1, value='Resumo por Tipo de Evento').font = SUBTITLE_FONT
    row += 1
    summary_headers = ['Tipo de Evento', 'Total Recebido', 'Nº Pagamentos']
    _apply_header(ws, row, summary_headers, [22, 18, 16])
    row += 1

    by_type = divs.groupby('event_type').agg(
        total=('total_income', 'sum'),
        payments=('num_payments', 'sum')
    ).reset_index()

    for _, r in by_type.iterrows():
        _apply_data_cell(ws, row, 1, str(r['event_type']), align='left')
        _apply_data_cell(ws, row, 2, r['total'], MONEY_FMT)
        _apply_data_cell(ws, row, 3, int(r['payments']), QTY_FMT)
        row += 1

    # Total
    for c in range(1, 4):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).font = TOTAL_FONT
    ws.cell(row=row, column=1, value='TOTAL')
    ws.cell(row=row, column=2, value=divs['total_income'].sum()).number_format = MONEY_FMT
    ws.cell(row=row, column=3, value=int(divs['num_payments'].sum())).number_format = QTY_FMT
    row += 2

    # ── Summary by ticker ──
    ws.cell(row=row, column=1, value='Proventos por Ticker').font = SUBTITLE_FONT
    row += 1
    ticker_headers = ['Ticker', 'Total Recebido', 'Nº Pagamentos']
    _apply_header(ws, row, ticker_headers, [14, 18, 16])
    row += 1

    by_ticker = divs.groupby('ticker').agg(
        total=('total_income', 'sum'),
        payments=('num_payments', 'sum')
    ).sort_values('total', ascending=False).reset_index()

    ticker_start = row
    for _, r in by_ticker.iterrows():
        _apply_data_cell(ws, row, 1, str(r['ticker']), align='left')
        _apply_data_cell(ws, row, 2, r['total'], MONEY_FMT)
        _apply_data_cell(ws, row, 3, int(r['payments']), QTY_FMT)
        row += 1

    # Bar chart
    bar = BarChart()
    bar.type = "col"
    bar.title = "Proventos por Ticker"
    bar.style = 10
    bar.width = 20
    bar.height = 12

    cats = Reference(ws, min_col=1, min_row=ticker_start, max_row=row - 1)
    vals = Reference(ws, min_col=2, min_row=ticker_start - 1, max_row=row - 1)
    bar.add_data(vals, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = "00B050"
    ws.add_chart(bar, f"E{ticker_start - 1}")

    row += 2

    # ── Full detail table ──
    ws.cell(row=row, column=1, value='Detalhamento Completo').font = SUBTITLE_FONT
    row += 1
    det_headers = ['Ticker', 'Período', 'Tipo Evento', 'Valor Total', 'Qtd Ações', 'Nº Pagamentos']
    _apply_header(ws, row, det_headers, [14, 12, 20, 16, 14, 16])
    row += 1
    det_start = row

    for _, r in divs.iterrows():
        _apply_data_cell(ws, row, 1, str(r['ticker']), align='left')
        _apply_data_cell(ws, row, 2, _period_to_label(r['_reference_period']), align='center')
        _apply_data_cell(ws, row, 3, str(r['event_type']), align='left')
        _apply_data_cell(ws, row, 4, r['total_income'], MONEY_FMT)
        _apply_data_cell(ws, row, 5, r['total_shares'], QTY_FMT)
        _apply_data_cell(ws, row, 6, int(r['num_payments']), QTY_FMT)
        row += 1

    ws.auto_filter.ref = f'A{det_start - 1}:F{row - 1}'


def build_negociacoes(wb: Workbook, data: Dict):
    """Sheet 5: Negociações (IR)"""
    ws = wb.create_sheet('Negociações')
    ws.sheet_properties.tabColor = 'FFC000'

    trades = data['trades']
    if trades.empty:
        return

    row = _add_title(ws, 1, 'Negociações', 'Compras e Vendas realizadas')

    headers = ['Ticker', 'Período', 'Instituição', 'Qtd Compra', 'Preço Médio Compra',
               'Total Compra', 'Qtd Venda', 'Preço Médio Venda', 'Total Venda']
    widths = [12, 12, 26, 14, 18, 16, 14, 18, 16]
    _apply_header(ws, row, headers, widths)
    row += 1
    data_start = row

    for _, t in trades.iterrows():
        _apply_data_cell(ws, row, 1, str(t['ticker']), align='left')
        _apply_data_cell(ws, row, 2, _period_to_label(t['_reference_period']), align='center')
        _apply_data_cell(ws, row, 3, str(t['institution']), align='left')

        buy_q = t['buy_quantity']
        _apply_data_cell(ws, row, 4, buy_q if buy_q > 0 else '-', QTY_FMT if buy_q > 0 else None)
        buy_p = t['avg_buy_price']
        _apply_data_cell(ws, row, 5, buy_p if buy_p > 0 else '-', MONEY_FMT if buy_p > 0 else None)
        buy_c = t['total_buy_cost']
        _apply_data_cell(ws, row, 6, buy_c if buy_c > 0 else '-', MONEY_FMT if buy_c > 0 else None)

        sell_q = t['sell_quantity']
        _apply_data_cell(ws, row, 7, sell_q if sell_q > 0 else '-', QTY_FMT if sell_q > 0 else None)
        sell_p = t['avg_sell_price']
        _apply_data_cell(ws, row, 8, sell_p if sell_p > 0 else '-', MONEY_FMT if sell_p > 0 else None)
        sell_v = t['total_sell_value']
        _apply_data_cell(ws, row, 9, sell_v if sell_v > 0 else '-', MONEY_FMT if sell_v > 0 else None)
        row += 1

    # Total row
    for c in range(1, 10):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).font = TOTAL_FONT
    ws.cell(row=row, column=1, value='TOTAL')
    for c, letter in [(4, 'D'), (6, 'F'), (7, 'G'), (9, 'I')]:
        ws.cell(row=row, column=c, value=f'=SUM({letter}{data_start}:{letter}{row - 1})')
        ws.cell(row=row, column=c).number_format = MONEY_FMT if c in (6, 9) else QTY_FMT

    ws.auto_filter.ref = f'A{data_start - 1}:I{row - 1}'
    ws.freeze_panes = f'A{data_start}'

    row += 2

    # ── Monthly buy/sell volume chart ──
    monthly_trades = trades.groupby('_reference_period').agg(
        total_compras=('total_buy_cost', 'sum'),
        total_vendas=('total_sell_value', 'sum')
    ).reset_index()

    ws.cell(row=row, column=1, value='Volume Mensal de Negociações').font = SUBTITLE_FONT
    row += 1
    vol_headers = ['Período', 'Total Compras', 'Total Vendas']
    _apply_header(ws, row, vol_headers, [12, 16, 16])
    row += 1
    vol_start = row

    for _, m in monthly_trades.iterrows():
        _apply_data_cell(ws, row, 1, _period_to_label(m['_reference_period']), align='center')
        _apply_data_cell(ws, row, 2, m['total_compras'], MONEY_FMT)
        _apply_data_cell(ws, row, 3, m['total_vendas'], MONEY_FMT)
        row += 1

    vol_end = row - 1

    bar = BarChart()
    bar.type = "col"
    bar.title = "Volume Mensal: Compras vs Vendas"
    bar.style = 10
    bar.width = 24
    bar.height = 12
    bar.y_axis.title = "R$"

    cats = Reference(ws, min_col=1, min_row=vol_start, max_row=vol_end)
    buy_vals = Reference(ws, min_col=2, min_row=vol_start - 1, max_row=vol_end)
    sell_vals = Reference(ws, min_col=3, min_row=vol_start - 1, max_row=vol_end)
    bar.add_data(buy_vals, titles_from_data=True)
    bar.add_data(sell_vals, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = "2E75B6"
    bar.series[1].graphicalProperties.solidFill = "FF6600"
    ws.add_chart(bar, f"E{vol_start - 1}")


def build_tesouro(wb: Workbook, data: Dict):
    """Sheet 6: Tesouro Direto"""
    ws = wb.create_sheet('Tesouro Direto')
    ws.sheet_properties.tabColor = '7030A0'

    treasury = data['treasury']
    if treasury.empty:
        ws.cell(row=1, column=1, value="Sem posições em Tesouro Direto").font = DATA_FONT
        return

    latest_period = treasury['_reference_period'].max()
    latest = treasury[treasury['_reference_period'] == latest_period]

    row = _add_title(ws, 1, 'Tesouro Direto', f'Posição em {_period_to_label(latest_period)}')

    headers = ['Produto', 'Indexador', 'Vencimento', 'Quantidade', 'Investido',
               'Valor Bruto', 'Valor Líquido', 'Valor Atual', 'Ganho/Perda']
    widths = [30, 12, 14, 12, 16, 16, 16, 16, 16]
    _apply_header(ws, row, headers, widths)
    row += 1
    data_start = row

    for _, t in latest.iterrows():
        _apply_data_cell(ws, row, 1, str(t['product_name']), align='left')
        _apply_data_cell(ws, row, 2, str(t['index_type']) if pd.notna(t['index_type']) else '', align='center')
        _apply_data_cell(ws, row, 3, str(t['maturity_date']) if pd.notna(t['maturity_date']) else '', align='center')
        _apply_data_cell(ws, row, 4, t['total_quantity'], QTY_FMT)
        _apply_data_cell(ws, row, 5, t['total_invested'], MONEY_FMT)
        _apply_data_cell(ws, row, 6, t['total_gross'], MONEY_FMT)
        _apply_data_cell(ws, row, 7, t['total_net'], MONEY_FMT)
        _apply_data_cell(ws, row, 8, t['total_current_value'], MONEY_FMT)
        gl_cell = _apply_data_cell(ws, row, 9, t['gain_loss'], MONEY_FMT)
        _apply_gain_loss_style(gl_cell, t['gain_loss'])
        row += 1

    # Total
    for c in range(1, 10):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).font = TOTAL_FONT
    ws.cell(row=row, column=1, value='TOTAL')
    for c, letter in [(5, 'E'), (6, 'F'), (7, 'G'), (8, 'H'), (9, 'I')]:
        ws.cell(row=row, column=c, value=f'=SUM({letter}{data_start}:{letter}{row - 1})')
        ws.cell(row=row, column=c).number_format = MONEY_FMT

    ws.auto_filter.ref = f'A{data_start - 1}:I{row - 1}'
    ws.freeze_panes = f'A{data_start}'

    row += 2

    # ── Monthly evolution chart ──
    monthly = treasury.groupby('_reference_period').agg(
        investido=('total_invested', 'sum'),
        atual=('total_current_value', 'sum')
    ).reset_index()

    ws.cell(row=row, column=1, value='Evolução Mensal').font = SUBTITLE_FONT
    row += 1
    evo_headers = ['Período', 'Investido', 'Valor Atual']
    _apply_header(ws, row, evo_headers, [12, 16, 16])
    row += 1
    evo_start = row

    for _, m in monthly.iterrows():
        _apply_data_cell(ws, row, 1, _period_to_label(m['_reference_period']), align='center')
        _apply_data_cell(ws, row, 2, m['investido'], MONEY_FMT)
        _apply_data_cell(ws, row, 3, m['atual'], MONEY_FMT)
        row += 1

    evo_end = row - 1

    chart = LineChart()
    chart.title = "Tesouro Direto: Investido vs Valor Atual"
    chart.style = 10
    chart.width = 24
    chart.height = 12
    chart.y_axis.title = "R$"

    cats = Reference(ws, min_col=1, min_row=evo_start, max_row=evo_end)
    for col_idx, color in [(2, "7030A0"), (3, "00B050")]:
        vals = Reference(ws, min_col=col_idx, min_row=evo_start - 1, max_row=evo_end)
        chart.add_data(vals, titles_from_data=True)
        chart.series[-1].graphicalProperties.line.width = 25000
        chart.series[-1].graphicalProperties.line.solidFill = color

    chart.set_categories(cats)
    ws.add_chart(chart, f"E{evo_start - 1}")


def build_renda_fixa(wb: Workbook, data: Dict):
    """Sheet 7: Renda Fixa"""
    ws = wb.create_sheet('Renda Fixa')
    ws.sheet_properties.tabColor = 'FFC000'

    fi = data['fixed_income']
    if fi.empty:
        ws.cell(row=1, column=1, value="Sem posições em Renda Fixa").font = DATA_FONT
        return

    latest_period = fi['_reference_period'].max()
    latest = fi[fi['_reference_period'] == latest_period]

    row = _add_title(ws, 1, 'Renda Fixa', f'CDB, LCI, LCA — Posição em {_period_to_label(latest_period)}')

    headers = ['Produto', 'Emissor', 'Código', 'Indexador', 'Regime',
               'Emissão', 'Vencimento', 'Quantidade', 'Valor Curva']
    widths = [20, 20, 14, 12, 12, 14, 14, 12, 16]
    _apply_header(ws, row, headers, widths)
    row += 1
    data_start = row

    for _, r in latest.iterrows():
        _apply_data_cell(ws, row, 1, str(r['product_name']) if pd.notna(r['product_name']) else '', align='left')
        _apply_data_cell(ws, row, 2, str(r['issuer']) if pd.notna(r['issuer']) else '', align='left')
        _apply_data_cell(ws, row, 3, str(r['code']) if pd.notna(r['code']) else '', align='center')
        _apply_data_cell(ws, row, 4, str(r['index_type']) if pd.notna(r['index_type']) else '', align='center')
        _apply_data_cell(ws, row, 5, str(r['regime_type']) if pd.notna(r['regime_type']) else '', align='center')
        _apply_data_cell(ws, row, 6, str(r['issue_date']) if pd.notna(r['issue_date']) else '', align='center')
        _apply_data_cell(ws, row, 7, str(r['maturity_date']) if pd.notna(r['maturity_date']) else '', align='center')
        _apply_data_cell(ws, row, 8, r['total_quantity'], QTY_FMT)
        _apply_data_cell(ws, row, 9, r['total_curva_value'], MONEY_FMT)
        row += 1

    # Total
    for c in range(1, 10):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).font = TOTAL_FONT
    ws.cell(row=row, column=1, value='TOTAL')
    ws.cell(row=row, column=9, value=f'=SUM(I{data_start}:I{row - 1})').number_format = MONEY_FMT

    ws.auto_filter.ref = f'A{data_start - 1}:I{row - 1}'
    ws.freeze_panes = f'A{data_start}'


def build_ir_resumo(wb: Workbook, data: Dict):
    """Sheet 8: Resumo para Declaração de IR"""
    ws = wb.create_sheet('IR - Resumo')
    ws.sheet_properties.tabColor = 'FF0000'

    row = _add_title(ws, 1, 'Resumo para Declaração de IR',
                     'Informações consolidadas para IRPF')

    # ── Bens e Direitos (positions at year end) ──
    ws.cell(row=row, column=1, value='Bens e Direitos — Ações e ETFs').font = SUBTITLE_FONT
    row += 1

    detail = data['stock_detail']
    if not detail.empty:
        latest_period = detail['_reference_period'].max()
        year = latest_period.split('-')[0]

        # Dec positions
        dec = detail[detail['_reference_period'] == latest_period]

        headers = ['Ticker', 'Tipo', 'Quantidade', 'Preço Médio', 'Custo Total (31/12)']
        widths = [12, 10, 12, 16, 18]
        _apply_header(ws, row, headers, widths)
        row += 1
        data_start = row

        for _, s in dec.iterrows():
            _apply_data_cell(ws, row, 1, str(s['ticker']), align='left')
            _apply_data_cell(ws, row, 2, 'ETF' if s['asset_type'] == 'etf' else 'Ação', align='center')
            _apply_data_cell(ws, row, 3, s['total_quantity'], QTY_FMT)
            _apply_data_cell(ws, row, 4, s['mean_purchase_price'], MONEY_FMT)
            _apply_data_cell(ws, row, 5, s['total_invested'], MONEY_FMT)
            row += 1

        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = TOTAL_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).font = TOTAL_FONT
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=5, value=f'=SUM(E{data_start}:E{row - 1})').number_format = MONEY_FMT
        row += 2

    # ── Rendimentos Isentos (Dividendos) ──
    divs = data['dividends']
    if not divs.empty:
        div_isentos = divs[divs['event_type'].str.contains('Dividendo', case=False, na=False)]
        if not div_isentos.empty:
            ws.cell(row=row, column=1, value='Rendimentos Isentos — Dividendos').font = SUBTITLE_FONT
            row += 1
            headers = ['Ticker', 'Total Recebido']
            _apply_header(ws, row, headers, [14, 18])
            row += 1
            div_start = row

            by_ticker = div_isentos.groupby('ticker')['total_income'].sum().sort_values(ascending=False)
            for ticker, total in by_ticker.items():
                _apply_data_cell(ws, row, 1, str(ticker), align='left')
                _apply_data_cell(ws, row, 2, total, MONEY_FMT)
                row += 1

            for c in range(1, 3):
                ws.cell(row=row, column=c).fill = TOTAL_FILL
                ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=c).font = TOTAL_FONT
            ws.cell(row=row, column=1, value='TOTAL')
            ws.cell(row=row, column=2, value=f'=SUM(B{div_start}:B{row - 1})').number_format = MONEY_FMT
            row += 2

        # ── Rendimentos Tributáveis (JCP) ──
        jcp = divs[divs['event_type'].str.contains('JCP|Juros', case=False, na=False)]
        if not jcp.empty:
            ws.cell(row=row, column=1, value='Rendimentos Tributáveis — JCP').font = SUBTITLE_FONT
            row += 1
            headers = ['Ticker', 'Total Recebido (Bruto)']
            _apply_header(ws, row, headers, [14, 22])
            row += 1
            jcp_start = row

            by_ticker = jcp.groupby('ticker')['total_income'].sum().sort_values(ascending=False)
            for ticker, total in by_ticker.items():
                _apply_data_cell(ws, row, 1, str(ticker), align='left')
                _apply_data_cell(ws, row, 2, total, MONEY_FMT)
                row += 1

            for c in range(1, 3):
                ws.cell(row=row, column=c).fill = TOTAL_FILL
                ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=c).font = TOTAL_FONT
            ws.cell(row=row, column=1, value='TOTAL')
            ws.cell(row=row, column=2, value=f'=SUM(B{jcp_start}:B{row - 1})').number_format = MONEY_FMT
            row += 2

    # ── Vendas realizadas (para apuração de ganho de capital) ──
    trades = data['trades']
    if not trades.empty:
        sells = trades[trades['sell_quantity'] > 0]
        if not sells.empty:
            ws.cell(row=row, column=1, value='Vendas Realizadas — Apuração de Ganho de Capital').font = SUBTITLE_FONT
            row += 1

            # Monthly sell totals (for R$20k exemption check)
            monthly_sells = sells.groupby('_reference_period')['total_sell_value'].sum().reset_index()

            headers = ['Período', 'Total Vendas', 'Acima R$20k?']
            _apply_header(ws, row, headers, [12, 18, 14])
            row += 1

            for _, m in monthly_sells.iterrows():
                _apply_data_cell(ws, row, 1, _period_to_label(m['_reference_period']), align='center')
                _apply_data_cell(ws, row, 2, m['total_sell_value'], MONEY_FMT)
                above = 'SIM' if m['total_sell_value'] > 20000 else 'NÃO'
                cell = _apply_data_cell(ws, row, 3, above, align='center')
                if above == 'SIM':
                    cell.font = Font(name='Arial', size=10, bold=True, color=DARK_RED)
                else:
                    cell.font = Font(name='Arial', size=10, color=DARK_GREEN)
                row += 1

            row += 1
            ws.cell(row=row, column=1,
                    value='* Vendas de ações abaixo de R$20.000/mês são isentas de IR (exceto day-trade e ETFs)').font = Font(
                name='Arial', size=9, italic=True, color='666666')


# ════════════════════════════════════════════════════════════════
# Main entry point
# ════════════════════════════════════════════════════════════════

def generate_report(duckdb_path: str, output_path: str):
    """Generate the complete Excel report."""
    log.info(f"Fetching data from {duckdb_path}")
    data = fetch_all_data(duckdb_path)

    periods = sorted(data['monthly_wallet']['_reference_period'].unique().tolist())
    log.info(f"Found {len(periods)} periods: {periods[0]} to {periods[-1]}")

    log.info("Fetching benchmark data (CDI, IBOV, DIVO11)")
    try:
        benchmarks = fetch_benchmarks(periods)
    except Exception as e:
        log.warning(f"Benchmark fetch failed: {e}. Using empty benchmarks.")
        benchmarks = pd.DataFrame({
            '_reference_period': periods,
            'cdi_monthly_return': 0, 'cdi_cumulative': 0,
            'ibov_return': 0, 'divo11_return': 0,
        })

    log.info("Building Excel report")
    wb = Workbook()

    build_resumo(wb, data, benchmarks)
    build_evolucao(wb, data, benchmarks)
    build_detalhamento_acoes(wb, data)
    build_proventos(wb, data)
    build_negociacoes(wb, data)
    build_tesouro(wb, data)
    build_renda_fixa(wb, data)
    build_ir_resumo(wb, data)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    log.info(f"Report saved to {output_path}")
    return output_path
